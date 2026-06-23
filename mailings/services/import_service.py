from __future__ import annotations

import logging
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from django.core.exceptions import ValidationError
from django.core.validators import validate_email
from django.db import IntegrityError, transaction
from django.utils import timezone
from openpyxl import load_workbook

from mailings.models import MailingRecord
from mailings.services.email_sender import send_email

logger = logging.getLogger(__name__)

REQUIRED_COLUMNS = ("external_id", "user_id", "email", "subject", "message")


@dataclass(frozen=True)
class ImportResult:
    processed: int = 0
    created: int = 0
    skipped: int = 0
    errors: int = 0


class MailingImportService:
    """Import mailing rows from an XLSX file and dispatch simulated emails."""

    def __init__(self, file_path: str | Path) -> None:
        self.file_path = Path(file_path)

    def run(self) -> ImportResult:
        if not self.file_path.exists():
            raise FileNotFoundError(f"File not found: {self.file_path}")

        workbook = load_workbook(self.file_path, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            rows = sheet.iter_rows(values_only=True)
            header_row = next(rows, None)
            if header_row is None:
                raise ValueError("XLSX file is empty.")

            column_map = self._build_column_map(header_row)
            result = ImportResult()

            for row_number, row in enumerate(rows, start=2):
                if self._is_empty_row(row):
                    continue

                result = ImportResult(
                    processed=result.processed + 1,
                    created=result.created,
                    skipped=result.skipped,
                    errors=result.errors,
                )

                try:
                    row_data = self._parse_row(row, column_map)
                    outcome = self._process_row(row_data)
                    if outcome == "created":
                        result = ImportResult(
                            processed=result.processed,
                            created=result.created + 1,
                            skipped=result.skipped,
                            errors=result.errors,
                        )
                    elif outcome == "skipped":
                        result = ImportResult(
                            processed=result.processed,
                            created=result.created,
                            skipped=result.skipped + 1,
                            errors=result.errors,
                        )
                except RowImportError as exc:
                    logger.warning(
                        "Row %s import error: %s",
                        row_number,
                        exc,
                    )
                    result = ImportResult(
                        processed=result.processed,
                        created=result.created,
                        skipped=result.skipped,
                        errors=result.errors + 1,
                    )

            return result
        finally:
            workbook.close()

    def _build_column_map(self, header_row: tuple[Any, ...]) -> dict[str, int]:
        headers = {
            str(value).strip().lower(): index
            for index, value in enumerate(header_row)
            if value is not None and str(value).strip()
        }

        missing = [column for column in REQUIRED_COLUMNS if column not in headers]
        if missing:
            raise ValueError(f"Missing required columns: {', '.join(missing)}")

        return headers

    def _parse_row(
        self,
        row: tuple[Any, ...],
        column_map: dict[str, int],
    ) -> dict[str, str]:
        data: dict[str, str] = {}

        for column in REQUIRED_COLUMNS:
            raw_value = row[column_map[column]] if column_map[column] < len(row) else None
            value = self._normalize_cell(raw_value)
            if not value:
                raise RowImportError(f"Column '{column}' is required.")
            data[column] = value

        try:
            validate_email(data["email"])
        except ValidationError as exc:
            raise RowImportError("; ".join(exc.messages)) from exc
        return data

    @staticmethod
    def _normalize_cell(value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).strip()

    @staticmethod
    def _is_empty_row(row: tuple[Any, ...]) -> bool:
        return all(
            cell is None or (isinstance(cell, str) and not cell.strip())
            for cell in row
        )

    def _process_row(self, row_data: dict[str, str]) -> str:
        external_id = row_data["external_id"]

        if MailingRecord.objects.filter(external_id=external_id).exists():
            logger.info("Skip duplicate external_id=%s", external_id)
            return "skipped"

        record = MailingRecord(
            external_id=external_id,
            user_id=row_data["user_id"],
            email=row_data["email"],
            subject=row_data["subject"],
            message=row_data["message"],
            status=MailingRecord.Status.SENT,
        )

        try:
            with transaction.atomic():
                record.full_clean()
                record.save()
        except ValidationError as exc:
            raise RowImportError(str(exc)) from exc
        except IntegrityError:
            logger.info("Skip duplicate external_id=%s (race condition)", external_id)
            return "skipped"

        try:
            send_email(
                email=record.email,
                subject=record.subject,
                message=record.message,
            )
            record.sent_at = timezone.now()
            record.save(update_fields=["sent_at"])
        except Exception as exc:
            record.status = MailingRecord.Status.FAILED
            record.error_message = str(exc)
            record.save(update_fields=["status", "error_message"])
            raise RowImportError(f"Failed to send email: {exc}") from exc

        return "created"


class RowImportError(Exception):
    """Raised when a single spreadsheet row cannot be imported."""
